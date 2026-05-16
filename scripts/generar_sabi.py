import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

METRICAS = [
    (15, 'ingresos'),
    (16, 'result_antes_imp'),
    (17, 'resultado_neto'),
    (18, 'activo_total'),
    (19, 'fondos_propios'),
    (22, 'roi_pct'),
    (23, 'roe_pct'),
    (24, 'liquidez'),
    (25, 'endeudamiento_pct'),
    (27, 'empleados'),
]
METRICAS_MONETARIAS = {'ingresos', 'result_antes_imp', 'resultado_neto', 'activo_total', 'fondos_propios'}

# Pernod 2019: resultado_neto y result_antes_imp tienen valores corruptos en SABI original
# (5.5B EUR para una empresa con 363M de ingresos). Se marcan como None.
PERNOD_FILAS_CORRUPTAS_2019 = {16, 17}

EMPRESAS = {
    'diageo':  ('Datos/SABI DIAGEO.xlsx',         'Diageo Espana SA',             'mil EUR', 'Madrid',    '30 jun', 'Principal'),
    'pernod':  ('Datos/SABI PERNOD RICARD.xlsx',  'Pernod Ricard Espana SA',      'EUR',     'Malaga',    '30 jun', 'Competidora'),
    'bacardi': ('Datos/SABI BACARDI.xlsx',         'Bacardi Espana SA',            'mil EUR', 'Barcelona', '31 mar', 'Competidora'),
    'beam':    ('Datos/SABI BEAM SUNTORY.xlsx',    'Beam Suntory Distribution SL', 'mil EUR', 'Madrid',    '31 dic', 'Competidora'),
}
ANOS_OBJETIVO = set(range(2019, 2025))


def mapear_fecha_valor(ws, fila_cabecera=10, fila_ref=15):
    cols_fecha = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=fila_cabecera, column=col).value
        if v is not None and '/' in str(v):
            cols_fecha[col] = v
    fechas_ord = sorted(cols_fecha.keys())
    resultado = {}
    for i, col_f in enumerate(fechas_ord):
        col_limite = fechas_ord[i - 1] if i > 0 else 0
        for col_v in range(col_f - 1, col_limite, -1):
            v = ws.cell(row=fila_ref, column=col_v).value
            if v is not None and isinstance(v, (int, float)):
                resultado[cols_fecha[col_f]] = col_v
                break
    return resultado


all_rows = []

for id_emp, (fichero, nombre, unidad, ciudad, cierre, tipo) in EMPRESAS.items():
    wb = openpyxl.load_workbook(fichero, data_only=True)
    ws = wb['Page 1']
    mapa = mapear_fecha_valor(ws)

    for fecha, col_v in sorted(mapa.items()):
        ano = int(str(fecha).split('/')[-1])
        if ano not in ANOS_OBJETIVO:
            continue

        row = {
            'id_empresa':    id_emp,
            'nombre':        nombre,
            'ciudad':        ciudad,
            'tipo':          tipo,
            'cierre_fiscal': cierre,
            'ano':           ano,
            'fecha_cierre':  fecha,
        }

        for fila, metrica in METRICAS:
            if id_emp == 'pernod' and ano == 2019 and fila in PERNOD_FILAS_CORRUPTAS_2019:
                row[metrica] = None
                continue

            valor = ws.cell(row=fila, column=col_v).value
            if isinstance(valor, str):
                valor = None

            if unidad == 'EUR' and metrica in METRICAS_MONETARIAS and valor is not None:
                valor = round(valor / 1000, 3)

            row[metrica] = valor

        ing = row.get('ingresos')
        rn = row.get('resultado_neto')
        emp = row.get('empleados')
        row['margen_neto_pct'] = round(rn / ing * 100, 3) if (ing and rn is not None) else None
        row['ingresos_por_empleado'] = round(ing / emp, 3) if (ing and emp) else None

        all_rows.append(row)

all_rows.sort(key=lambda r: (r['id_empresa'], r['ano']))

wb_out = openpyxl.Workbook()

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F4E79')
alt_fill = PatternFill('solid', fgColor='D6E4F0')

# ── Hoja fact_kpis ─────────────────────────────────────────────────────────────

ws1 = wb_out.active
ws1.title = 'fact_kpis'

COLS_F1 = [
    'id_empresa', 'nombre', 'ciudad', 'tipo', 'cierre_fiscal', 'ano', 'fecha_cierre',
    'ingresos', 'result_antes_imp', 'resultado_neto', 'activo_total', 'fondos_propios',
    'roi_pct', 'roe_pct', 'liquidez', 'endeudamiento_pct', 'empleados',
    'margen_neto_pct', 'ingresos_por_empleado',
]
CABECERA_F1 = [
    'id_empresa', 'nombre', 'ciudad', 'tipo', 'cierre_fiscal', 'ano', 'fecha_cierre',
    'ingresos (mil EUR)', 'result_antes_imp (mil EUR)', 'resultado_neto (mil EUR)',
    'activo_total (mil EUR)', 'fondos_propios (mil EUR)',
    'roi_pct (%)', 'roe_pct (%)', 'liquidez', 'endeudamiento_pct (%)',
    'empleados', 'margen_neto_pct (%)', 'ingresos_por_empleado (mil EUR/emp)',
]

for col_idx, cab in enumerate(CABECERA_F1, 1):
    cell = ws1.cell(row=1, column=col_idx, value=cab)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for row_idx, r in enumerate(all_rows, 2):
    fill = alt_fill if row_idx % 2 == 0 else None
    for col_idx, campo in enumerate(COLS_F1, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=r.get(campo))
        if fill:
            cell.fill = fill

anchos1 = [12, 22, 12, 12, 12, 6, 12, 16, 20, 20, 18, 18, 10, 10, 10, 20, 10, 18, 24]
for i, w in enumerate(anchos1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'
ws1.row_dimensions[1].height = 40

# ── Hoja Datos originales ──────────────────────────────────────────────────────

ws2 = wb_out.create_sheet('Datos originales')

COLS_F2 = [
    'id_empresa', 'nombre', 'ciudad', 'tipo', 'cierre_fiscal', 'ano', 'fecha_cierre',
    'ingresos', 'result_antes_imp', 'resultado_neto', 'activo_total', 'fondos_propios',
    'roi_pct', 'roe_pct', 'liquidez', 'endeudamiento_pct', 'empleados',
]

for col_idx, cab in enumerate(COLS_F2, 1):
    cell = ws2.cell(row=1, column=col_idx, value=cab)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

for row_idx, r in enumerate(all_rows, 2):
    fill = alt_fill if row_idx % 2 == 0 else None
    for col_idx, campo in enumerate(COLS_F2, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=r.get(campo))
        if fill:
            cell.fill = fill

anchos2 = [12, 22, 12, 12, 12, 6, 12, 14, 18, 18, 16, 18, 10, 10, 10, 18, 10]
for i, w in enumerate(anchos2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

ruta = 'Datos-Tratados/SABI.xlsx'
wb_out.save(ruta)
print(f'Guardado: {ruta}')
print(f'Filas: {len(all_rows)} (4 empresas x 6 anos)')
print()

por_emp = defaultdict(list)
for r in all_rows:
    por_emp[r['id_empresa']].append(r)

for emp, rows in sorted(por_emp.items()):
    print(f'{emp}:')
    for r in rows:
        ing = r.get('ingresos')
        margen = r.get('margen_neto_pct')
        i_emp = r.get('ingresos_por_empleado')
        print(f"  {r['ano']} ({r['fecha_cierre']}): ingresos={ing}, margen={margen}%, ing/emp={i_emp}")
    print()
